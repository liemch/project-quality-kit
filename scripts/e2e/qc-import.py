#!/usr/bin/env python3
"""Fallback QC Excel import when npm `xlsx` is unavailable. Writes qc/catalog.json."""
from __future__ import annotations

import argparse
import json
import os
import re
import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    import subprocess

    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl", "-q"])
    import openpyxl

KIT_ROOT = Path(__file__).resolve().parents[2]
SKIP = {
    "Cover",
    "Guideline",
    "Revision History",
    "Summary",
    "Dashboard",
    "Report Test",
    "Bug Data",
    "RTM",
}


def refuse_base() -> None:
    if KIT_ROOT.name == "project-quality-kit" and os.environ.get("QUALITY_ALLOW_BASE_INIT") != "1":
        print(
            f"[qc:import.py] Refusing to mutate Base template '{KIT_ROOT}'. "
            "Clone to <project>-quality first. Maintainer override: QUALITY_ALLOW_BASE_INIT=1.",
            file=sys.stderr,
        )
        sys.exit(1)


def load_cfg():
    p = KIT_ROOT / "_meta" / "project.json"
    if p.exists():
        return json.loads(p.read_text(encoding="utf-8"))
    yml = KIT_ROOT / "_meta" / "project.yml"
    if yml.exists():
        try:
            import yaml  # type: ignore

            return yaml.safe_load(yml.read_text(encoding="utf-8")) or {}
        except Exception:
            return {}
    return {}


def norm_header(cell) -> str:
    """Normalize Excel header: newlines / pipes / trailing (hint) → stable key."""
    s = str(cell or "")
    s = s.replace("\n", " ").replace("|", " ")
    s = re.sub(r"\s+", " ", s).strip().lower()
    s = re.sub(r"\s*\([^)]*\)\s*$", "", s).strip()
    return s


def col(fmap, row, *names):
    for n in names:
        i = fmap.get(n.lower())
        if i is not None and i < len(row) and row[i] is not None:
            return str(row[i]).strip()
    return ""


def main():
    refuse_base()
    ap = argparse.ArgumentParser()
    ap.add_argument("--file", default="")
    ap.add_argument("--priority", default="")
    ap.add_argument("--group", default="")
    ap.add_argument("--sheet", default="")
    args = ap.parse_args()

    cfg = load_cfg()
    file = args.file
    if not file:
        input_dir = KIT_ROOT / "qc" / "input"
        files = sorted(input_dir.glob("*.xlsx")) if input_dir.exists() else []
        files = [f for f in files if not f.name.startswith("~$")]
        if not files:
            print("[qc:import.py] Provide --file or place xlsx in qc/input/", file=sys.stderr)
            sys.exit(1)
        file = str(files[0])
    path = Path(file).resolve()
    if not path.exists():
        print("[qc:import.py] File not found:", path, file=sys.stderr)
        sys.exit(1)

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    cases = []
    for sheet_name in wb.sheetnames:
        if sheet_name in SKIP:
            continue
        if args.sheet and sheet_name != args.sheet:
            continue
        rows = list(wb[sheet_name].iter_rows(values_only=True))
        header_idx = None
        fmap = {}
        for i, row in enumerate(rows[:20]):
            cells = [str(c) if c is not None else "" for c in (row or ())]
            if any(re.search(r"testcase\s*id", c, re.I) for c in cells):
                header_idx = i
                for idx, c in enumerate(row or ()):
                    key = norm_header(c)
                    if key:
                        fmap[key] = idx
                break
        if header_idx is None:
            continue
        for row in rows[header_idx + 1 :]:
            row = row or ()
            id_ = col(fmap, row, "testcase id")
            if not id_ or not id_.startswith("TC_"):
                continue
            priority = col(fmap, row, "priority")
            group = col(fmap, row, "group")
            if args.priority and args.priority.lower() not in priority.lower():
                continue
            if args.group and group.lower() != args.group.lower():
                continue
            cases.append(
                {
                    "id": id_,
                    "reqId": col(fmap, row, "req id"),
                    "docSource": col(fmap, row, "doc source"),
                    "group": group,
                    "priority": priority,
                    "title": col(fmap, row, "test title"),
                    "precondition": col(fmap, row, "pre-condition"),
                    "steps": col(fmap, row, "test steps"),
                    "expected": col(fmap, row, "expected result"),
                    "origin": col(fmap, row, "origin"),
                    "sheet": sheet_name,
                    "automated": col(fmap, row, "automated"),
                    "script": col(fmap, row, "script"),
                }
            )
    wb.close()

    out_rel = (cfg.get("qc") or {}).get("catalog_out") or "qc/catalog.json"
    out = KIT_ROOT / out_rel
    out.parent.mkdir(parents=True, exist_ok=True)
    from datetime import datetime, timezone

    payload = {
        "source": str(path),
        "importedAt": datetime.now(timezone.utc).isoformat().replace("+00:00", "Z"),
        "count": len(cases),
        "cases": cases,
    }
    out.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"[qc:import.py] {len(cases)} case(s) → {out}")
    from collections import Counter

    print("[qc:import.py] by priority:", dict(Counter(c["priority"] for c in cases)))
    print("[qc:import.py] by group:", dict(Counter(c["group"] for c in cases)))


if __name__ == "__main__":
    main()
